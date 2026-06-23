"""One-off: read the two duplicate-tracking XLSX files Clint downloaded
from Shrewsbury's GHL export, attach a recommended_action per row, and
emit a single styled cleanup-plan XLSX the school's CFO can review.

Output: ../shrewsbury-duplicates-cleanup-plan.xlsx
"""

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

opps_dup = pd.read_excel(
    r'C:\Users\thelo\Downloads\Duplicates in Opportunities.xlsx',
    sheet_name='Duplicates ',
)
contacts = pd.read_excel(
    r'C:\Users\thelo\Downloads\Duplicates in ALL CONTACTS.xlsx',
    sheet_name='Sheet1',
)


# Detect sibling-style clusters: shared Contact ID OR shared email but
# different opportunity names. Sibling opps shouldn't be auto-deleted —
# they represent separate kids in the pipeline.
def opp_classify(cluster_df):
    by_contact = cluster_df['Contact ID'].nunique()
    by_oppname = cluster_df['Opportunity Name'].str.lower().str.strip().nunique()
    by_email = cluster_df['email'].nunique()
    is_same_family = by_contact == 1 or by_email == 1
    has_different_names = by_oppname > 1
    if is_same_family and has_different_names:
        return 'REVIEW — possible siblings'
    return 'safe to follow suggest_keep'


opp_rows = []
for cid, g in opps_dup.groupby('name_duplicate_cluster'):
    classification = opp_classify(g)
    for _, r in g.iterrows():
        opp_rows.append({
            'cluster': cid,
            'cluster_size': len(g),
            'classification': classification,
            'recommended_action': (
                'KEEP — survivor of cluster' if r['suggest_keep']
                else (
                    'REVIEW — possible sibling, do not auto-delete'
                    if 'REVIEW' in classification
                    else 'DELETE — duplicate'
                )
            ),
            'Opportunity Name': r['Opportunity Name'],
            'Contact Name': r['Contact Name'],
            'stage': r['stage'],
            'email': r['email'],
            'phone': r['phone'],
            'Updated on': str(r['Updated on'])[:10] if pd.notna(r['Updated on']) else '',
            'tags': str(r['tags'])[:80] if pd.notna(r['tags']) else '',
            'Opportunity ID': r['Opportunity ID'],
            'Contact ID': r['Contact ID'],
            'completeness_score': r['completeness_score'],
        })
opp_df = pd.DataFrame(opp_rows).sort_values(['cluster', 'recommended_action'])

contacts['name_key'] = (
    contacts['First Name'].str.strip().str.lower() + '|'
    + contacts['Last Name'].str.strip().str.lower()
)
contact_rows = []
for key, g in contacts.groupby('name_key'):
    has_data = g['Email'].notna() | g['Phone'].notna()
    n_data = has_data.sum()
    n_total = len(g)
    for _, r in g.sort_values('Original Row #').iterrows():
        row_has_data = pd.notna(r['Email']) or pd.notna(r['Phone'])
        if n_data == 0:
            action = 'REVIEW — every row in cluster is blank'
        elif row_has_data:
            action = 'KEEP — populated row'
        else:
            action = 'DELETE — blank shell'
        contact_rows.append({
            'cluster_name': f"{r['First Name']} {r['Last Name']}",
            'cluster_size': n_total,
            'recommended_action': action,
            'Original Row #': int(r['Original Row #']) if pd.notna(r['Original Row #']) else None,
            'First Name': r['First Name'],
            'Last Name': r['Last Name'],
            'Email': r['Email'] if pd.notna(r['Email']) else '',
            'Phone': r['Phone'] if pd.notna(r['Phone']) else '',
            'Reason': r['Reason'],
        })
contact_df = pd.DataFrame(contact_rows).sort_values(['cluster_name', 'recommended_action'])

opp_counts = opp_df['recommended_action'].value_counts()
ct_counts = contact_df['recommended_action'].value_counts()

print("OPPORTUNITIES recommended actions:")
for k, v in opp_counts.items():
    print(f"  {v:3} | {k}")
print("\nCONTACTS recommended actions:")
for k, v in ct_counts.items():
    print(f"  {v:3} | {k}")

out_path = r'C:\Users\thelo\OneDrive\Desktop\Growth Suite - NEW\shrewsbury-duplicates-cleanup-plan.xlsx'
wb = Workbook()
del wb[wb.sheetnames[0]]

ws = wb.create_sheet('Summary')
ws['A1'] = 'Shrewsbury Montessori — Duplicate Cleanup Plan'
ws['A1'].font = Font(size=14, bold=True)
ws['A3'] = 'Source files'
ws['A3'].font = Font(bold=True)
ws['A4'] = 'Duplicates in Opportunities.xlsx'
ws['A5'] = 'Duplicates in ALL CONTACTS.xlsx'

ws['A7'] = 'Opportunities recommended actions'
ws['A7'].font = Font(bold=True)
r = 8
for k, v in opp_counts.items():
    ws.cell(row=r, column=1, value=k)
    ws.cell(row=r, column=2, value=int(v))
    r += 1
ws.cell(row=r, column=1, value='Total').font = Font(bold=True)
ws.cell(row=r, column=2, value=int(opp_counts.sum())).font = Font(bold=True)
r += 2

ws.cell(row=r, column=1, value='Contacts recommended actions').font = Font(bold=True)
r += 1
for k, v in ct_counts.items():
    ws.cell(row=r, column=1, value=k)
    ws.cell(row=r, column=2, value=int(v))
    r += 1
ws.cell(row=r, column=1, value='Total').font = Font(bold=True)
ws.cell(row=r, column=2, value=int(ct_counts.sum())).font = Font(bold=True)
r += 2

ws.cell(row=r, column=1, value='How to use this file').font = Font(bold=True)
r += 1
notes = [
    '1. Open the Opportunities and Contacts tabs. Each row carries a recommended_action.',
    '2. Sort by cluster (already sorted) so duplicate groups appear together.',
    '3. KEEP (green) rows are survivors — leave them as-is in Growth Suite.',
    '4. DELETE (red) rows are confirmed duplicates: empty shells / rejected candidates.',
    '5. REVIEW (amber) rows are clusters that may be siblings sharing one parent contact — DO NOT auto-delete. Inspect names + stages and decide manually.',
    '6. Once Clint or Sonia signs off, send the file back; I will execute the deletes via the Growth Suite API (no manual clicking in GHL).',
]
for n in notes:
    ws.cell(row=r, column=1, value=n)
    r += 1
ws.column_dimensions['A'].width = 80
ws.column_dimensions['B'].width = 12


def write_df(ws, df):
    header_fill = PatternFill('solid', start_color='1F4E78')
    header_font = Font(bold=True, color='FFFFFF')
    keep_fill = PatternFill('solid', start_color='C6EFCE')
    delete_fill = PatternFill('solid', start_color='FFC7CE')
    review_fill = PatternFill('solid', start_color='FFEB9C')
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=1, column=j, value=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(vertical='center', horizontal='left')
    action_col = df.columns.get_loc('recommended_action') + 1
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        for j, col in enumerate(df.columns, start=1):
            v = row[col]
            if pd.isna(v):
                v = ''
            ws.cell(row=i, column=j, value=v)
        action_val = str(row['recommended_action'])
        if action_val.startswith('KEEP'):
            fill = keep_fill
        elif action_val.startswith('DELETE'):
            fill = delete_fill
        else:
            fill = review_fill
        ws.cell(row=i, column=action_col).fill = fill
        ws.cell(row=i, column=action_col).font = Font(bold=True)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
    for j, col in enumerate(df.columns, start=1):
        sample_max = df[col].astype(str).str.len().max() if len(df) else 0
        max_len = max(len(str(col)), int(sample_max))
        ws.column_dimensions[get_column_letter(j)].width = min(max(max_len + 2, 12), 50)


write_df(wb.create_sheet('Opportunities'), opp_df)
write_df(wb.create_sheet('Contacts'), contact_df)
wb.save(out_path)
print(f"\nWrote: {out_path}")
