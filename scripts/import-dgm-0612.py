#!/usr/bin/env python3
"""
DGM full fresh import (2026-06-12) — THE new source of truth:

  1. "Sonia Reviewed DB.xlsx"          -> students (277; registrar-reviewed)
  2. "Parent Dump 6 11.xlsx"           -> parents/families (569 rows)
  3. "FACTS 2026-27 Transactions.xlsx" -> facts_transactions ledger (260)

Match key across all three: students.metadata->>'unique_id'
(= Sonia UNIQUE ID = dump UNIQUE ID = FACTS Student ID, sometimes with
a "-1" suffix for split-household second ledgers).

Operator decision: these files OVERWRITE current data.
  - Students: matched by unique_id -> every sheet-mapped column/metadata
    key is overwritten (sheet-mapped keys are cleared first, so stale
    values from earlier imports/GHL can't linger). Non-sheet keys
    (ghl_contact_id, ghl_slot, apid, program, form_completion) survive.
  - Status: ENROLLMENT STAT drives both students.status
    (Withdrawn -> 'withdrawn', else 'active') and the 2026-27
    enrollment row (enrolled/accepted/pending/withdrawn).
  - Parents: sheet-is-truth per family (same rules as the 0611 import:
    match by email, update, insert missing, deactivate extras;
    ghl_contact_id/logins/PINs preserved on matched parents).
  - Active students NOT in Sonia -> archived (status='inactive');
    families left with no active students archived + parents deactivated.
  - FACTS rows upserted into facts_transactions (unique per
    school+unique_id+year). Unmatched "-1" ids link via base id.

The Carter (DEMO) family is excluded from everything.

  python import-dgm-0612.py            # dry-run
  python import-dgm-0612.py --apply
"""
import sys, re, json, datetime
import openpyxl
import psycopg2
import psycopg2.extras

SONIA = r"C:\Users\thelo\Downloads\Sonia Reviewed DB.xlsx"
DUMP  = r"C:\Users\thelo\Downloads\Parent Dump 6 11.xlsx"
FACTS = r"C:\Users\thelo\Downloads\FACTS 2026-27 Transactions.xlsx"
SCHOOL_ID = "cfa9030d-c8fe-49ae-a9e7-f1003844ec07"
DEMO_FAMILY_ID = "cdf70975-b0a4-4f3a-8a34-2858bfffe750"
YEAR = "2026-27"
APPLY = "--apply" in sys.argv

def db_url():
    for line in open(".env.local", encoding="utf-8"):
        m = re.match(r"^DATABASE_URL=(.*)$", line.strip())
        if m: return m.group(1).strip().strip('"')
    raise SystemExit("DATABASE_URL not found")

def s(v):
    if v is None: return None
    t = str(v).strip()
    return t or None

def norm_email(v):
    t = s(v)
    return t.lower() if t and "@" in t else None

def norm_phone(v):
    if v is None: return None
    d = re.sub(r"\D", "", str(v))
    if len(d) == 10: d = "1" + d
    return d or None

def iso_dt(v):
    if v is None: return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return datetime.datetime(v.year, v.month, v.day).strftime("%Y-%m-%dT00:00:00.000Z")
    return s(v)

def as_date(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return datetime.date(v.year, v.month, v.day)
    return None

def cents(v):
    t = s(v)
    if t is None: return 0
    t = t.replace(",", "").replace("$", "")
    try:
        return round(float(t) * 100)
    except ValueError:
        return 0

# ── Sonia sheet ─────────────────────────────────────────────────────
wb = openpyxl.load_workbook(SONIA, read_only=True, data_only=True)
sonia_rows = [r for r in list(wb.active.iter_rows(values_only=True))[1:] if r and s(r[4])]
wb.close()

ENR_STATUS = {"Enrolled": "enrolled", "Accepted": "accepted", "Pending": "pending", "Withdrawn": "withdrawn"}

# Metadata keys this import OWNS: cleared from previous metadata before
# the sheet values land, so stale values can't survive under these keys.
def sonia_metadata(r, prev=None):
    md = dict(prev or {})
    choice = s(r[30])
    mapped = {
        "unique_id": s(r[4]), "household_id": s(r[5]),
        "first_name": s(r[3]), "last_name": s(r[2]),
        "language": s(r[6]), "grade_level": s(r[7]), "age": s(r[8]),
        "student_street": s(r[9]), "student_city": s(r[10]),
        "student_state": s(r[11]), "student_zip": s(r[12]),
        "homeroom": s(r[13]), "birth_date": iso_dt(r[15]),
        "lead_teacher": s(r[16]), "daily_schedule": s(r[17]),
        "program_name": s(r[18]), "initial_start_date": iso_dt(r[19]),
        "withdrawal_date": iso_dt(r[20]), "ethnicity": s(r[21]),
        "non_custodial_restriction": s(r[22]), "restriction_description": s(r[23]),
        "graduation_year": s(r[24]), "withdrawal_fee": s(r[25]),
        "emergency_card_received": iso_dt(r[26]),
        "payment_plan": s(r[27]), "enrollment_start_date": iso_dt(r[28]),
        "extended_day": s(r[29]),
        # program_tuition feeds the roster Tuition column ($ extraction);
        # when the choice string has no $ the column falls back to
        # tuition_fee, so only set it when it carries a price.
        "program_tuition": choice if (choice and "$" in choice) else None,
        "program_tuition_choice": choice,
        "referred_by": s(r[31]), "organic_lunch": s(r[32]),
        "months_enrolled": s(r[33]), "tuition_fee": s(r[34]),
        "admin_fee_percentage": s(r[35]), "discount_type": s(r[36]),
        "discount_percentage": s(r[37]), "financial_aid": s(r[38]),
        "field_needed_for_tuition": s(r[39]),
        "health_care_provider": s(r[40]), "health_care_provider_phone": s(r[41]),
        "emergency_first_contact": s(r[42]), "allergy": s(r[43]),
        "do_not_remove_person": s(r[44]), "do_not_remove_person_2": s(r[45]),
        "custody_paperwork": s(r[46]),
        "five04_plan": s(r[47]), "504_plan": s(r[47]), "iep": s(r[48]),
        "legal_authority": s(r[49]), "legal_authority_other": s(r[50]),
        "physical_custody": s(r[51]), "physical_custody_other": s(r[52]),
        "enrollment_status": s(r[0]), "re_enroll_form_submitted": iso_dt(r[1]),
    }
    for k in mapped:
        md.pop(k, None)
    for k, v in mapped.items():
        if v is not None:
            md[k] = v
    return md

# ── Parent dump ─────────────────────────────────────────────────────
wb2 = openpyxl.load_workbook(DUMP, read_only=True, data_only=True)
dump_rows = [r for r in list(wb2.active.iter_rows(values_only=True))[1:] if r and s(r[14])]
wb2.close()

dump_students = {}   # uid -> {apid, hh, program, parents:[...]}
for r in dump_rows:
    uid = s(r[14])
    st = dump_students.setdefault(uid, {"apid": s(r[1]), "hh": s(r[4]), "program": s(r[17]), "parents": []})
    em = norm_email(r[7])
    p = {"first": s(r[5]) or "", "last": s(r[6]) or "", "email": em, "phone": norm_phone(r[13])}
    if not any(x["email"] == em and x["first"] == p["first"] for x in st["parents"]):
        st["parents"].append(p)

# ── FACTS ───────────────────────────────────────────────────────────
wb3 = openpyxl.load_workbook(FACTS, read_only=True, data_only=True)
facts_rows = [r for r in list(wb3.active.iter_rows(values_only=True))[1:] if r and s(r[2])]
wb3.close()

CHARGE_COLS = [
    (3, "annual_tuition"), (4, "administrative_fee"), (5, "late_fee"),
    (6, "organic_lunch"), (7, "extended_day"), (8, "late_pickup_fee"),
    (9, "not_signed_out_fee"), (10, "enrollment_fee"), (11, "enrichment"),
    (12, "athletics"), (13, "withdrawal_fee"), (14, "sst_tuition"),
    (15, "change_fee"), (16, "chromebook_fee"), (17, "childcare"),
    (18, "hearing_vision"),
]
CREDIT_COLS = [
    (20, "annual_discount"), (21, "sibling_discount"),
    (22, "employee_discount"), (23, "financial_aid"), (24, "miscellaneous"),
]

print(f"=== DGM 0612 import {'(APPLY)' if APPLY else '(DRY-RUN)'} ===")
print(f"Sonia students: {len(sonia_rows)} · dump parent-rows: {len(dump_rows)} ({len(dump_students)} students) · FACTS ledgers: {len(facts_rows)}\n")

conn = psycopg2.connect(db_url())
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

# ── DB state ────────────────────────────────────────────────────────
cur.execute("""SELECT s.id, s.family_id, s.status, s.first_name, s.last_name, s.metadata
               FROM students s WHERE s.school_id=%s""", (SCHOOL_ID,))
db_students = cur.fetchall()
by_uid = {}
hh_to_family = {}
for x in db_students:
    md = x["metadata"] or {}
    uid = md.get("unique_id")
    if uid: by_uid[str(uid)] = x
    hh = md.get("household_id")
    if hh and hh not in hh_to_family:
        hh_to_family[hh] = x["family_id"]

cur.execute("SELECT id, name FROM classrooms WHERE school_id=%s", (SCHOOL_ID,))
classroom_by_name = {x["name"].strip().lower(): x["id"] for x in cur.fetchall() if x["name"]}

# ── Phase 0: missing classrooms (Suite 100, Tower) ──────────────────
sheet_homerooms = {s(r[13]) for r in sonia_rows if s(r[13])}
missing_rooms = [h for h in sorted(sheet_homerooms) if h.lower() not in classroom_by_name]
for h in missing_rooms:
    if APPLY:
        cur.execute("INSERT INTO classrooms (school_id, name, academic_year) VALUES (%s,%s,%s) RETURNING id", (SCHOOL_ID, h, YEAR))
        classroom_by_name[h.lower()] = cur.fetchone()["id"]
print(f"phase 0: {'created' if APPLY else 'would create'} {len(missing_rooms)} missing classrooms: {missing_rooms}")

# ── Phase 1: students from Sonia ────────────────────────────────────
sonia_uids = set()
new_fam_cache = {}
stats = {"updated": 0, "created": 0, "fam_created": 0, "withdrawn": 0,
         "enr_updated": 0, "enr_inserted": 0}
uid_to_family = {}

for r in sonia_rows:
    uid = s(r[4]); sonia_uids.add(uid)
    stat_raw = s(r[0]) or "Enrolled"
    enr_status = ENR_STATUS.get(stat_raw, "enrolled")
    stu_status = "withdrawn" if stat_raw == "Withdrawn" else "active"
    if stu_status == "withdrawn": stats["withdrawn"] += 1
    first, last = s(r[3]), s(r[2])
    dob, gender = as_date(r[15]), s(r[14])
    hh = s(r[5])
    homeroom = s(r[13])
    classroom_id = classroom_by_name.get((homeroom or "").lower())
    schedule = s(r[17])
    allergy = s(r[43])

    ex = by_uid.get(uid)
    if ex:
        sid, fam_id = ex["id"], ex["family_id"]
        new_md = sonia_metadata(r, ex["metadata"])
        if APPLY:
            cur.execute("""UPDATE students SET first_name=%s, last_name=%s,
                             date_of_birth=COALESCE(%s,date_of_birth), gender=COALESCE(%s,gender),
                             status=%s, metadata=%s::jsonb, updated_at=now()
                           WHERE id=%s""",
                        (first, last, dob, gender, stu_status, json.dumps(new_md), sid))
        stats["updated"] += 1
    else:
        fam_id = hh_to_family.get(hh) or new_fam_cache.get(hh)
        if fam_id is None:
            stats["fam_created"] += 1
            if APPLY:
                cur.execute("INSERT INTO families (school_id, display_name, status) VALUES (%s,%s,'active') RETURNING id",
                            (SCHOOL_ID, f"{last} Family"))
                fam_id = cur.fetchone()["id"]
            else:
                fam_id = f"<new:{hh}>"
            if hh: new_fam_cache[hh] = fam_id
        md = sonia_metadata(r)
        sid = None
        if APPLY:
            cur.execute("""INSERT INTO students (family_id, school_id, first_name, last_name, date_of_birth, gender, status, metadata)
                           VALUES (%s,%s,%s,%s,%s,%s,%s,%s::jsonb) RETURNING id""",
                        (fam_id, SCHOOL_ID, first, last, dob, gender, stu_status, json.dumps(md)))
            sid = cur.fetchone()["id"]
        stats["created"] += 1
    uid_to_family[uid] = fam_id

    # 2026-27 enrollment upsert (other years untouched)
    if APPLY and sid:
        cur.execute("""UPDATE enrollments SET status=%s, classroom_id=%s, schedule=%s, updated_at=now()
                       WHERE student_id=%s AND academic_year=%s RETURNING id""",
                    (enr_status, classroom_id, schedule, sid, YEAR))
        if cur.fetchone():
            stats["enr_updated"] += 1
        else:
            cur.execute("""INSERT INTO enrollments (student_id, school_id, classroom_id, academic_year, status, schedule, enrolled_at, metadata)
                           VALUES (%s,%s,%s,%s,%s,%s,now(),'{}'::jsonb)""",
                        (sid, SCHOOL_ID, classroom_id, YEAR, enr_status, schedule))
            stats["enr_inserted"] += 1
        if allergy:
            cur.execute("""INSERT INTO student_health_profiles (school_id, student_id, allergies)
                           VALUES (%s,%s,%s)
                           ON CONFLICT (school_id, student_id) DO UPDATE SET allergies=EXCLUDED.allergies, updated_at=now()""",
                        (SCHOOL_ID, sid, allergy))

print(f"phase 1: matched/updated {stats['updated']} · created {stats['created']} students "
      f"({stats['fam_created']} new families) · withdrawn {stats['withdrawn']} · "
      f"enrollments {stats['enr_updated']} updated / {stats['enr_inserted']} inserted")

# ── Phase 2: parents from dump (sheet is truth) + per-student program ─
fam_parents = {}
prog_updates = 0
for uid, st in dump_students.items():
    fam = uid_to_family.get(uid)
    if not fam or fam == DEMO_FAMILY_ID: continue
    lst = fam_parents.setdefault(fam, [])
    for p in st["parents"]:
        if p["email"] and not any(x["email"] == p["email"] for x in lst):
            lst.append(p)
        elif not p["email"] and not any(x["first"] == p["first"] and x["last"] == p["last"] for x in lst):
            lst.append(p)
    # numbered program (##PROGRAM) + apid live only in the dump
    if APPLY:
        patch = {}
        if st["program"]: patch["program"] = st["program"]
        if st["apid"]: patch["apid"] = st["apid"]
        if patch:
            cur.execute("""UPDATE students SET metadata = metadata || %s::jsonb, updated_at=now()
                           WHERE school_id=%s AND metadata->>'unique_id'=%s""",
                        (json.dumps(patch), SCHOOL_ID, uid))
            prog_updates += cur.rowcount

upd = ins = deact = 0
for fam, plist in fam_parents.items():
    if str(fam).startswith("<new:"):
        ins += len(plist)
        continue
    cur.execute("""SELECT id, first_name, last_name, lower(email) email, status
                   FROM parents WHERE family_id=%s AND school_id=%s""", (fam, SCHOOL_ID))
    existing = cur.fetchall()
    by_email = {x["email"]: x for x in existing if x["email"]}
    sheet_emails = {p["email"] for p in plist if p["email"]}
    for p in plist:
        match = by_email.get(p["email"]) if p["email"] else None
        if match:
            if APPLY:
                cur.execute("""UPDATE parents SET first_name=%s, last_name=%s,
                                 phone=COALESCE(%s, phone), status='active', updated_at=now()
                               WHERE id=%s""", (p["first"], p["last"], p["phone"], match["id"]))
            upd += 1
        else:
            if APPLY:
                cur.execute("""INSERT INTO parents (family_id, school_id, first_name, last_name, email, phone, role, is_primary, status)
                               VALUES (%s,%s,%s,%s,%s,%s,'parent',false,'active')""",
                            (fam, SCHOOL_ID, p["first"], p["last"], p["email"], p["phone"]))
            ins += 1
    for x in existing:
        if x["status"] == "active" and (x["email"] or "") not in sheet_emails:
            if APPLY:
                cur.execute("UPDATE parents SET status='inactive', is_primary=false, updated_at=now() WHERE id=%s", (x["id"],))
            deact += 1
print(f"phase 2: parents — update {upd}, insert {ins}, deactivate {deact} "
      f"(across {len(fam_parents)} families) · program/apid patched on {prog_updates} students")

# ── Phase 3: archive non-sheet active students + empty families ─────
to_archive = [x for x in db_students
              if x["status"] == "active" and x["family_id"] != DEMO_FAMILY_ID
              and str((x["metadata"] or {}).get("unique_id") or "") not in sonia_uids]
if APPLY and to_archive:
    cur.execute("UPDATE students SET status='inactive', updated_at=now() WHERE id = ANY(%s::uuid[])",
                ([str(x["id"]) for x in to_archive],))
print(f"phase 3: {'archived' if APPLY else 'would archive'} {len(to_archive)} active students not in Sonia: "
      f"{[(x['first_name'] + ' ' + x['last_name']) for x in to_archive][:10]}")
if APPLY:
    cur.execute("""UPDATE families f SET status='inactive', updated_at=now()
                   WHERE f.school_id=%s AND f.id <> %s AND f.status='active'
                     AND NOT EXISTS (SELECT 1 FROM students s WHERE s.family_id=f.id AND s.status='active')
                   RETURNING f.id""", (SCHOOL_ID, DEMO_FAMILY_ID))
    fam_arch = [x["id"] for x in cur.fetchall()]
    if fam_arch:
        cur.execute("UPDATE parents SET status='inactive', is_primary=false, updated_at=now() WHERE family_id = ANY(%s::uuid[])",
                    ([str(f) for f in fam_arch],))
    print(f"  archived {len(fam_arch)} now-empty families (+ parents)")

# one primary parent per active family
if APPLY:
    cur.execute("""
      WITH fams AS (SELECT id FROM families WHERE school_id=%s AND status='active'),
      ranked AS (
        SELECT p.id, ROW_NUMBER() OVER (PARTITION BY p.family_id ORDER BY p.is_primary DESC, p.created_at) rn
        FROM parents p JOIN fams f ON f.id=p.family_id WHERE p.status='active'
      )
      UPDATE parents p SET is_primary = (r.rn = 1)
      FROM ranked r WHERE r.id = p.id AND (p.is_primary <> (r.rn=1))""", (SCHOOL_ID,))
    print(f"  primary flags normalized ({cur.rowcount} adjusted)")

# ── Phase 4: FACTS ledger ───────────────────────────────────────────
fx = {"linked": 0, "via_base": 0, "unlinked": 0, "upserts": 0}
if APPLY:
    cur.execute("""SELECT s.metadata->>'unique_id' uid, s.id FROM students s
                   WHERE s.school_id=%s AND s.metadata ? 'unique_id'""", (SCHOOL_ID,))
    uid_to_sid = {x["uid"]: x["id"] for x in cur.fetchall()}
else:
    uid_to_sid = {u: f"<sid:{u}>" for u in by_uid} | {u: f"<sid:{u}>" for u in sonia_uids}

for r in facts_rows:
    uid = s(r[2])
    sid = uid_to_sid.get(uid)
    if sid:
        fx["linked"] += 1
    else:
        base = re.sub(r"-\d+$", "", uid)
        sid = uid_to_sid.get(base)
        if sid: fx["via_base"] += 1
        else: fx["unlinked"] += 1
    charges = {k: cents(r[i]) for i, k in CHARGE_COLS if cents(r[i]) != 0}
    credits = {k: cents(r[i]) for i, k in CREDIT_COLS if cents(r[i]) != 0}
    if APPLY:
        cur.execute("""INSERT INTO facts_transactions
                         (school_id, student_id, unique_id, academic_year, parent_name, student_name,
                          charges, credits, total_charges_cents, total_credits_cents, net_charges_cents,
                          payments_cents, credits_applied_cents, remaining_balance_cents, source_file)
                       VALUES (%s,%s,%s,%s,%s,%s,%s::jsonb,%s::jsonb,%s,%s,%s,%s,%s,%s,%s)
                       ON CONFLICT (school_id, unique_id, academic_year) DO UPDATE SET
                         student_id=EXCLUDED.student_id, parent_name=EXCLUDED.parent_name,
                         student_name=EXCLUDED.student_name, charges=EXCLUDED.charges,
                         credits=EXCLUDED.credits, total_charges_cents=EXCLUDED.total_charges_cents,
                         total_credits_cents=EXCLUDED.total_credits_cents, net_charges_cents=EXCLUDED.net_charges_cents,
                         payments_cents=EXCLUDED.payments_cents, credits_applied_cents=EXCLUDED.credits_applied_cents,
                         remaining_balance_cents=EXCLUDED.remaining_balance_cents,
                         source_file=EXCLUDED.source_file, imported_at=now()""",
                    (SCHOOL_ID, sid if not str(sid).startswith("<") else None, uid, YEAR,
                     s(r[0]), s(r[1]), json.dumps(charges), json.dumps(credits),
                     cents(r[19]), cents(r[25]), cents(r[26]), cents(r[27]), cents(r[28]), cents(r[29]),
                     "FACTS 2026-27 Transactions.xlsx"))
        fx["upserts"] += 1
print(f"phase 4: FACTS — {fx['linked']} linked, {fx['via_base']} via base id, {fx['unlinked']} unlinked"
      + (f" · {fx['upserts']} upserted" if APPLY else " (write at apply)"))

# ── Spot-checks ─────────────────────────────────────────────────────
for r in sonia_rows:
    if s(r[2]) == "Alami":
        print(f"  spot: {s(r[3])} Alami — stat={s(r[0])} homeroom={s(r[13])} lead={s(r[16])} "
              f"tuition_choice={s(r[30])!r} gross={s(r[34])}")

if APPLY:
    conn.commit()
    cur.execute("""SELECT COUNT(*) FILTER (WHERE status='active')::int act,
                          COUNT(*) FILTER (WHERE status='withdrawn')::int wd,
                          COUNT(*)::int total FROM students WHERE school_id=%s""", (SCHOOL_ID,))
    x = cur.fetchone()
    cur.execute("SELECT COUNT(*)::int n FROM parents WHERE school_id=%s AND status='active' AND email IS NOT NULL", (SCHOOL_ID,))
    pe = cur.fetchone()["n"]
    cur.execute("SELECT COUNT(*)::int n, COUNT(student_id)::int linked FROM facts_transactions WHERE school_id=%s", (SCHOOL_ID,))
    ft = cur.fetchone()
    print(f"\nCOMMITTED. students: {x['act']} active / {x['wd']} withdrawn / {x['total']} total · "
          f"active parents w/ email: {pe} · facts rows: {ft['n']} ({ft['linked']} linked)")
else:
    conn.rollback()
    print("\nDRY-RUN — nothing written. Re-run with --apply.")
cur.close(); conn.close()
